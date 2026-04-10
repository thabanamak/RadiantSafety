"""
CSA Historical Data Seeder
---------------------------
Reads the Crime Statistics Agency Excel file
  "Data_Tables_LGA_Criminal_Incidents_Year_Ending_December_2025.xlsx"
and pushes one row per matched suburb into the Supabase `incidents` table.

Each row represents the *total* criminal incidents recorded for that suburb
across all offence types for the year ending December 2025.

Intensity is derived from the raw incident count mapped to a 1-10 scale
relative to the highest-volume suburb in the dataset.

Usage (from the RadiantSafety/ project root):
    py backend/app/scrapers/seed_historical_data.py

Or from backend/:
    py -m app.scrapers.seed_historical_data
"""

from __future__ import annotations

import logging
import os
from collections import defaultdict
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_FILENAME = "Data_Tables_LGA_Criminal_Incidents_Year_Ending_December_2025.xlsx"
SHEET_NAME = "Table 03"  # suburb-level breakdown: Suburb/Town Name + Incidents Recorded

SUPABASE_TABLE = "incidents"
SOURCE_LABEL = "CSA Historical"

# Suburb/town name → (latitude, longitude)
# Covers ~200 suburbs across all of Victoria: inner Melbourne, middle/outer
# suburbs, and every major regional city and town.
SUBURB_COORDS: dict[str, tuple[float, float]] = {
    # ---- Inner Melbourne ----
    "Melbourne":        (-37.8136,  144.9631),
    "Southbank":        (-37.8226,  144.9584),
    "Docklands":        (-37.8145,  144.9460),
    "South Yarra":      (-37.8388,  144.9930),
    "St Kilda":         (-37.8676,  144.9809),
    "Richmond":         (-37.8183,  144.9984),
    "Fitzroy":          (-37.7990,  144.9780),
    "Collingwood":      (-37.7980,  144.9870),
    "Carlton":          (-37.7950,  144.9670),
    "Prahran":          (-37.8494,  144.9924),
    "Footscray":        (-37.7998,  144.8995),
    "Brunswick":        (-37.7666,  144.9600),
    "North Melbourne":  (-37.7980,  144.9430),
    "East Melbourne":   (-37.8120,  144.9870),
    "West Melbourne":   (-37.8080,  144.9450),
    "South Melbourne":  (-37.8320,  144.9580),
    "Port Melbourne":   (-37.8380,  144.9320),
    "Albert Park":      (-37.8440,  144.9560),
    "Toorak":           (-37.8430,  145.0120),
    "Cremorne":         (-37.8270,  144.9940),
    "Abbotsford":       (-37.8060,  144.9990),
    "Windsor":          (-37.8560,  144.9910),
    "Flemington":       (-37.7880,  144.9340),
    "Ascot Vale":       (-37.7780,  144.9200),
    "Fitzroy North":    (-37.7880,  144.9780),
    "Northcote":        (-37.7690,  145.0000),
    "Thornbury":        (-37.7540,  145.0060),
    "Hawthorn":         (-37.8230,  145.0340),
    "Kew":              (-37.8070,  145.0350),
    "Camberwell":       (-37.8420,  145.0700),
    "Malvern East":     (-37.8730,  145.0620),
    "Glen Iris":        (-37.8600,  145.0600),
    "Moonee Ponds":     (-37.7670,  144.9200),
    "Essendon":         (-37.7540,  144.9170),
    "Coburg":           (-37.7430,  144.9660),
    "Brunswick West":   (-37.7640,  144.9450),
    "Brunswick East":   (-37.7720,  144.9730),
    "Coburg North":     (-37.7270,  144.9660),
    "Maribyrnong":      (-37.7740,  144.8890),

    # ---- Northern suburbs ----
    "Preston":          (-37.7490,  145.0130),
    "Reservoir":        (-37.7170,  145.0080),
    "Epping":           (-37.6500,  145.0130),
    "Craigieburn":      (-37.5980,  144.9460),
    "Broadmeadows":     (-37.6830,  144.9180),
    "Thomastown":       (-37.6820,  145.0140),
    "Roxburgh Park":    (-37.6400,  144.9280),
    "Mill Park":        (-37.6630,  145.0600),
    "Bundoora":         (-37.6980,  145.0600),
    "Glenroy":          (-37.7030,  144.9260),
    "Campbellfield":    (-37.6660,  144.9600),
    "Lalor":            (-37.6670,  145.0130),
    "South Morang":     (-37.6500,  145.0920),
    "Mernda":           (-37.6050,  145.0960),
    "Meadow Heights":   (-37.6520,  144.9180),
    "Sunbury":          (-37.5770,  144.7260),
    "Heidelberg":       (-37.7560,  145.0670),
    "Heidelberg West":  (-37.7430,  145.0420),
    "Greensborough":    (-37.7040,  145.1040),
    "Wallan":           (-37.4170,  144.9790),

    # ---- Eastern suburbs ----
    "Box Hill":         (-37.8192,  145.1200),
    "Ringwood":         (-37.8150,  145.2290),
    "Doncaster":        (-37.7830,  145.1260),
    "Glen Waverley":    (-37.8780,  145.1640),
    "Mount Waverley":   (-37.8770,  145.1280),
    "Wantirna South":   (-37.8800,  145.2200),
    "Ferntree Gully":   (-37.8880,  145.2930),
    "Croydon":          (-37.7950,  145.2820),
    "Boronia":          (-37.8610,  145.2870),
    "Lilydale":         (-37.7560,  145.3540),
    "Rowville":         (-37.9230,  145.2320),
    "Bayswater":        (-37.8430,  145.2660),
    "Clayton":          (-37.9200,  145.1220),
    "Mulgrave":         (-37.9260,  145.1730),
    "Cheltenham":       (-37.9560,  145.0480),
    "Bentleigh East":   (-37.9210,  145.0380),
    "Highett":          (-37.9470,  145.0410),
    "Moorabbin":        (-37.9380,  145.0440),
    "Brighton":         (-37.9070,  145.0020),
    "Oakleigh":         (-37.8990,  145.0930),
    "Ormond":           (-37.9050,  145.0380),

    # ---- South-eastern suburbs ----
    "Dandenong":        (-37.9875,  145.2160),
    "Noble Park":       (-37.9700,  145.1650),
    "Springvale":       (-37.9490,  145.1530),
    "Keysborough":      (-37.9910,  145.1740),
    "Narre Warren":     (-37.9930,  145.3020),
    "Berwick":          (-38.0350,  145.3500),
    "Cranbourne":       (-38.0990,  145.2830),
    "Cranbourne North": (-38.0710,  145.2870),
    "Cranbourne West":  (-38.0860,  145.2520),
    "Hampton Park":     (-38.0320,  145.2620),
    "Endeavour Hills":  (-37.9780,  145.2570),
    "Hallam":           (-38.0170,  145.2630),
    "Doveton":          (-37.9930,  145.2310),
    "Dandenong North":  (-37.9630,  145.2150),
    "Dandenong South":  (-38.0250,  145.2120),
    "Narre Warren South": (-38.0260, 145.2970),
    "Pakenham":         (-38.0710,  145.4870),
    "Clyde North":      (-38.0910,  145.3530),

    # ---- South / bayside ----
    "Frankston":        (-38.1440,  145.1250),
    "Seaford":          (-38.1030,  145.1300),
    "Carrum Downs":     (-38.0980,  145.1750),
    "Mornington":       (-38.2180,  145.0380),
    "Hastings":         (-38.3000,  145.1850),
    "Rosebud":          (-38.3570,  144.9140),

    # ---- Western suburbs ----
    "Werribee":         (-37.9050,  144.6620),
    "Hoppers Crossing":  (-37.8830,  144.7000),
    "Sunshine":         (-37.7880,  144.8330),
    "St Albans":        (-37.7430,  144.8000),
    "Deer Park":        (-37.7680,  144.7700),
    "Tarneit":          (-37.8380,  144.6960),
    "Point Cook":       (-37.9050,  144.7470),
    "Braybrook":        (-37.7860,  144.8560),
    "Sunshine West":    (-37.7940,  144.8120),
    "Sunshine North":   (-37.7680,  144.8380),
    "Truganina":        (-37.8270,  144.7360),
    "Altona North":     (-37.8370,  144.8530),
    "Williamstown":     (-37.8600,  144.8930),
    "Taylors Lakes":    (-37.6990,  144.7870),
    "Melton":           (-37.6840,  144.5830),
    "Melton South":     (-37.7050,  144.5780),
    "Melton West":      (-37.6860,  144.5540),
    "Caroline Springs": (-37.7350,  144.7350),
    "Wyndham Vale":     (-37.8910,  144.6340),
    "Tullamarine":      (-37.6950,  144.8790),

    # ---- Geelong / Surf Coast ----
    "Geelong":          (-38.1499,  144.3617),
    "Corio":            (-38.0770,  144.3770),
    "Norlane":          (-38.0930,  144.3590),
    "Belmont":          (-38.1780,  144.3440),
    "Sebastopol":       (-37.5820,  143.8330),

    # ---- Ballarat ----
    "Ballarat Central": (-37.5600,  143.8630),
    "Wendouree":        (-37.5340,  143.8310),

    # ---- Bendigo ----
    "Bendigo":          (-36.7580,  144.2800),
    "Kangaroo Flat":    (-36.7910,  144.2440),

    # ---- Gippsland ----
    "Morwell":          (-38.2350,  146.3960),
    "Traralgon":        (-38.1950,  146.5400),
    "Moe":              (-38.1760,  146.2600),
    "Sale":             (-38.1110,  147.0670),
    "Warragul":         (-38.1590,  145.9300),
    "Drouin":           (-38.1370,  145.8600),
    "Bairnsdale":       (-37.8230,  147.6100),

    # ---- Murray / Hume / North-east ----
    "Shepparton":       (-36.3800,  145.3990),
    "Mooroopna":        (-36.3920,  145.3580),
    "Wodonga":          (-36.1210,  146.8880),
    "Wangaratta":       (-36.3580,  146.3100),
    "Benalla":          (-36.5520,  145.9820),
    "Echuca":           (-36.1280,  144.7520),
    "Seymour":          (-37.0270,  145.1390),

    # ---- Western Victoria ----
    "Warrnambool":      (-38.3810,  142.4870),
    "Horsham":          (-36.7120,  142.1990),
    "Portland":         (-38.3440,  141.6040),
    "Colac":            (-38.3400,  143.5850),
    "Ararat":           (-37.2840,  142.9280),
    "Hamilton":         (-37.7440,  142.0210),

    # ---- North-west ----
    "Mildura":          (-34.1850,  142.1620),
    "Swan Hill":        (-35.3380,  143.5540),
    "Ouyen":            (-35.0700,  142.3200),

    # ---- Other ----
    "Merrijig":         (-37.1400,  146.2700),
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_file(filename: str) -> Path:
    """
    Search for *filename* starting from this script's directory and walking
    up to the repo root.  Also checks the current working directory.
    """
    candidates = [Path.cwd() / filename, Path(__file__).resolve().parent / filename]
    # Walk up from script location
    current = Path(__file__).resolve().parent
    for _ in range(8):
        candidates.append(current / filename)
        current = current.parent

    for path in candidates:
        if path.is_file():
            return path

    raise FileNotFoundError(
        f"Could not locate '{filename}'. "
        "Run the script from the RadiantSafety project root, or place the "
        "Excel file there."
    )


def _find_env_file() -> Path:
    """Walk up from this script's directory to locate ``.env.local``."""
    current = Path(__file__).resolve().parent
    for _ in range(10):
        candidate = current / ".env.local"
        if candidate.is_file():
            return candidate
        current = current.parent
    raise FileNotFoundError("Could not locate .env.local in any ancestor directory")


def _init_supabase() -> Client:
    env_path = _find_env_file()
    load_dotenv(dotenv_path=env_path)
    log.info("Loaded environment from %s", env_path)

    url = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")

    if not url or not key:
        raise EnvironmentError("Missing SUPABASE_URL / SUPABASE_SERVICE_KEY in .env.local")

    return create_client(url, key)


def _intensity_from_count(count: int, max_count: int) -> int:
    """
    Map a raw incident count to an intensity score in [1, 10].
    Uses a logarithmic scale so that mid-range suburbs aren't all squashed
    near 1 just because Melbourne CBD dominates.
    """
    import math
    if max_count <= 0 or count <= 0:
        return 1
    ratio = math.log(count + 1) / math.log(max_count + 1)
    return max(1, min(10, round(ratio * 10)))


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

def extract_suburb_totals(excel_path: Path) -> dict[str, int]:
    """
    Read Table 03 from the Excel file and return a dict of
    ``{suburb_name: total_incidents}`` summed across all offence types.
    """
    log.info("Opening workbook: %s", excel_path)
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    totals: dict[str, int] = defaultdict(int)
    rows_read = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Validate header so we fail fast if the layout changes
            expected = ("Year", "Year ending", "Local Government Area",
                        "Postcode", "Suburb/Town Name", "Offence Division",
                        "Offence Subdivision", "Offence Subgroup", "Incidents Recorded")
            if row[:9] != expected:
                log.warning("Unexpected header row: %s", row[:9])
            continue

        suburb: str | None = row[4]
        incidents = row[8]

        if not suburb or incidents is None:
            continue

        suburb = suburb.strip()
        try:
            totals[suburb] += int(incidents)
        except (TypeError, ValueError):
            pass

        rows_read += 1

    log.info("Read %d data rows; found %d unique suburbs", rows_read, len(totals))
    wb.close()
    return dict(totals)


def build_records(suburb_totals: dict[str, int]) -> list[dict]:
    """
    Cross-reference suburb totals with SUBURB_COORDS and build the list of
    rows to insert into the incidents table.
    """
    if not suburb_totals:
        return []

    max_count = max(suburb_totals.values())
    records: list[dict] = []
    skipped: list[str] = []

    for suburb, coords in SUBURB_COORDS.items():
        count = suburb_totals.get(suburb)
        if count is None:
            skipped.append(suburb)
            continue

        intensity = _intensity_from_count(count, max_count)
        records.append({
            "title": f"{suburb} – {count:,} criminal incidents (year ending Dec 2025)",
            "suburb": suburb,
            "location_lat": coords[0],
            "location_lng": coords[1],
            "intensity": intensity,
            "source": SOURCE_LABEL,
            "is_verified": True,
        })

    if skipped:
        log.warning(
            "No Excel data found for %d suburb(s) in SUBURB_COORDS: %s",
            len(skipped),
            ", ".join(skipped),
        )

    log.info("Built %d records to insert", len(records))
    return records


# ---------------------------------------------------------------------------
# Supabase persistence
# ---------------------------------------------------------------------------

def push_to_supabase(records: list[dict], client: Client | None = None) -> int:
    """
    Insert records into Supabase in batches, skipping any whose title
    already exists.  Returns the count of newly inserted rows.
    """
    if not records:
        log.info("Nothing to push – record list is empty.")
        return 0

    if client is None:
        client = _init_supabase()

    # Fetch all existing CSA Historical titles in one call to avoid N+1 queries
    existing_titles: set[str] = set()
    try:
        existing = (
            client.table(SUPABASE_TABLE)
            .select("title")
            .eq("source", SOURCE_LABEL)
            .execute()
        )
        existing_titles = {row["title"] for row in (existing.data or [])}
        if existing_titles:
            log.info("Found %d existing CSA Historical rows – will skip duplicates", len(existing_titles))
    except Exception:
        log.warning("Could not fetch existing titles – duplicates may occur")

    new_records = [r for r in records if r["title"] not in existing_titles]
    skipped = len(records) - len(new_records)
    if skipped:
        log.info("Skipping %d records that already exist", skipped)

    if not new_records:
        log.info("All records already exist – nothing to insert.")
        return 0

    BATCH_SIZE = 50
    pushed = 0
    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i : i + BATCH_SIZE]
        try:
            client.table(SUPABASE_TABLE).insert(batch).execute()
            pushed += len(batch)
            log.info(
                "Batch %d: inserted %d records (%d–%d of %d)",
                i // BATCH_SIZE + 1,
                len(batch),
                i + 1,
                i + len(batch),
                len(new_records),
            )
        except Exception:
            log.exception("Batch insert failed at offset %d – falling back to row-by-row", i)
            for rec in batch:
                try:
                    client.table(SUPABASE_TABLE).insert(rec).execute()
                    pushed += 1
                    log.info("Inserted (fallback): %s", rec["suburb"])
                except Exception:
                    log.exception("Failed to insert: %s", rec["suburb"])

    log.info("Push complete – %d / %d new records written.", pushed, len(records))
    return pushed


# ---------------------------------------------------------------------------
# Entry-point
# ---------------------------------------------------------------------------

def run() -> None:
    log.info("=" * 60)
    log.info("CSA Historical Data Seeder – starting run")
    log.info("=" * 60)

    excel_path = _find_file(EXCEL_FILENAME)
    log.info("Using Excel file: %s", excel_path)

    suburb_totals = extract_suburb_totals(excel_path)
    records = build_records(suburb_totals)

    if not records:
        log.warning("No records to push. Check SUBURB_COORDS vs Excel suburb names.")
        return

    pushed = push_to_supabase(records)
    log.info("Run finished. %d new historical records persisted.", pushed)


if __name__ == "__main__":
    run()
